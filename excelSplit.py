from openpyxl import Workbook, load_workbook

originalFileLocation = "/home/gaon/Downloads/Police.csv"

Ex = load_workbook(filename=originalFileLocation)

data_folder = ("/data")


sheet1 = Ex['Sheet']

Exs = []
tmp1=0
tmp2=6

areaC = []
areaD = []


for i in sheet1.rows:
    if i[0].value == tmp1 and i[1].value == tmp2:

        areaC.append(i[2].value)
        areaD.append(i[3].value)


    else:
        folderName = saveFolder + "/" + str(tmp1) + str(tmp2) +".xlsx"
        folder = open(folderName, "w")
        folder.close()

        wb = Workbook()
        sheet2 = wb.active
        sheet2.title = "sheet"

        for j in range(len(areaC)):
            sheet2.cell(row=j+1,column=1).value=tmp1
            sheet2.cell(row=j+1, column=2).value=tmp2
            sheet2.cell(row=j+1,column=3).value=areaC[j]
            sheet2.cell(row=j+1,column=4).value=areaD[j]
        wb.save(filename=folderName)

        areaC = []
        areaD = []
        tmp1 = i[0].value
        tmp2 = i[1].value

        areaC.append(i[2].value)
        areaD.append(i[3].value)
